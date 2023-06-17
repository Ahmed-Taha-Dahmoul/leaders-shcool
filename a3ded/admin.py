from django.contrib import admin
from django import forms
from django.http import HttpResponse
from openpyxl import Workbook 
from .models import Teacher, SchoolClass, Grade, Subject, Student, Trimestre, Arbi, Francai, Anglai


class SubjectInline(admin.StackedInline):
    model = Subject


class StudentInline(admin.StackedInline):
    model = Student


class SubjectAdmin(admin.ModelAdmin):
    list_display = ('name', 'school_class')

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if obj and obj.school_class.english == 2:
            form.base_fields['anglai_oral'] = forms.DecimalField()
        return form

    def get_fields(self, request, obj=None):
        fields = super().get_fields(request, obj)
        if obj and obj.school_class.english == 2:
            fields.append('anglai_oral')
        return fields

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        trimestres = Trimestre.objects.all()
        if obj.name.lower() == 'arbi':
            students = obj.school_class.student_set.all()

            for trimestre in trimestres:
                for student in students:
                    Arbi.objects.get_or_create(
                        trimestre=trimestre,
                        student=student,
                        subject=obj,
                        defaults={
                            'test': 0.0,
                            'chifehi': 0.0,
                            'kira2a': 0.0,
                            'intej': 0.0,
                        }
                    )
        elif obj.name.lower() == 'francai':
            students = obj.school_class.student_set.all()

            for trimestre in trimestres:
                for student in students:
                    Francai.objects.get_or_create(
                        trimestre=trimestre,
                        student=student,
                        subject=obj,
                        defaults={
                            'test': 0.0,
                            'exp': 0.0,
                            'lecture': 0.0,
                            'production': 0.0,
                        }
                    )
        elif obj.name.lower() == 'anglai':
            students = obj.school_class.student_set.all()

            for trimestre in trimestres:
                for student in students:
                    Anglai.objects.get_or_create(
                        trimestre=trimestre,
                        student=student,
                        subject=obj,
                        defaults={
                            'test': 0.0,
                            'oral': 0.0,
                        }
                    )
        else:
            students = obj.school_class.student_set.all()

            for trimestre in trimestres:
                for student in students:
                    Grade.objects.get_or_create(
                        trimestre=trimestre,
                        student=student,
                        subject=obj,
                        defaults={
                            'test': 0.0,
                        }
                    )





from openpyxl import Workbook
from django.http import HttpResponse
from .models import Grade, Francai, Arbi, Anglai, Trimestre, SchoolClass
from openpyxl.styles import Alignment


from openpyxl.utils import get_column_letter



from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.http import HttpResponse

import csv


class SchoolClassAdmin(admin.ModelAdmin):
    filter_horizontal = ('teachers',)
    inlines = [SubjectInline, StudentInline]
    list_display = ('name',)

    def get_inline_instances(self, request, obj=None):
        if not obj:
            return []
        return super().get_inline_instances(request, obj)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)

        class_instance = form.instance
        subjects = class_instance.subject_set.all()

        for subject in subjects:
            trimestres = Trimestre.objects.all()
            students = subject.school_class.student_set.all()

            if subject.name.lower() == 'arbi':
                for trimestre in trimestres:
                    for student in students:
                        Arbi.objects.get_or_create(
                            trimestre=trimestre,
                            student=student,
                            subject=subject,
                            defaults={
                                'test': 0.0,
                                'chifehi': 0.0,
                                'kira2a': 0.0,
                                'intej': 0.0,
                            }
                        )
            elif subject.name.lower() == 'francai':
                for trimestre in trimestres:
                    for student in students:
                        Francai.objects.get_or_create(
                            trimestre=trimestre,
                            student=student,
                            subject=subject,
                            defaults={
                                'test': 0.0,
                                'exp': 0.0,
                                'lecture': 0.0,
                                'production': 0.0,
                            }
                        )
            elif subject.name.lower() == 'anglai':
                for trimestre in trimestres:
                    for student in students:
                        Anglai.objects.get_or_create(
                            trimestre=trimestre,
                            student=student,
                            subject=subject,
                            defaults={
                                'test': 0.0,
                                'oral': 0.0,
                            }
                        )
            else:
                for trimestre in trimestres:
                    for student in students:
                        Grade.objects.get_or_create(
                            trimestre=trimestre,
                            student=student,
                            subject=subject,
                            defaults={
                                'test': 0.0,
                            }
                        )

    def export_student_grades(self, request, queryset):
        trimestres = Trimestre.objects.all()
        subjects = Subject.objects.filter(school_class__in=queryset)
        students = Student.objects.filter(school_class__in=queryset)

        workbook = Workbook()

        for trimestre in trimestres:
            sheet = workbook.create_sheet(title=str(trimestre))

            # Write header row
            header_row = ['Student']
            for subject in subjects:
                header_row.append(subject.name)
                if subject.name.lower() == 'anglai' and subject.school_class.english == 2:
                    header_row.append('Oral')
                if subject.name.lower() == 'arbi' :
                    header_row.append('Chifehi')
                    header_row.append('Kira2a')
                    header_row.append('Intej')
                if subject.name.lower() == 'francai' :
                    header_row.append('Exp')
                    header_row.append('Lecture')
                    header_row.append('Production')
            sheet.append(header_row)

            # Write student data
            for student in students:
                data_row = [student.name]
                for subject in subjects:
                    try:
                        grade = Grade.objects.get(trimestre=trimestre, student=student, subject=subject)
                        data_row.append(grade.test)
                        if subject.name.lower() == 'anglai' and subject.school_class.english == 2:
                            try:
                                anglai_grade = Anglai.objects.get(trimestre=trimestre, student=student, subject=subject)
                                data_row.append(anglai_grade.oral)
                            except Anglai.DoesNotExist:
                                data_row.append('-')
                        if subject.name.lower() == 'arbi' :
                            try:
                                arbi_grade = Arbi.objects.get(trimestre=trimestre, student=student, subject=subject)
                                data_row.append(arbi_grade.chifehi)
                                data_row.append(arbi_grade.kira2a)
                                data_row.append(arbi_grade.intej)
                            except Arbi.DoesNotExist:
                                data_row.extend(['-', '-', '-'])
                        if subject.name.lower() == 'francai' :
                            try:
                                francai_grade = Francai.objects.get(trimestre=trimestre, student=student, subject=subject)
                                data_row.append(francai_grade.exp)
                                data_row.append(francai_grade.lecture)
                                data_row.append(francai_grade.production)
                            except Francai.DoesNotExist:
                                data_row.extend(['-', '-', '-'])
                    except Grade.DoesNotExist:
                        data_row.append('-')
                        if subject.name.lower() == 'anglai' and subject.school_class.english == 2:
                            data_row.append('-')
                        if subject.name.lower() == 'arbi' :
                            data_row.extend(['-', '-', '-'])
                        if subject.name.lower() == 'francai' :
                            data_row.extend(['-', '-', '-'])
                sheet.append(data_row)

            # Apply alignment to cells
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        # Remove the default sheet created by openpyxl
        workbook.remove(workbook.active)

        # Create HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(subject.school_class.name)

        # Save the workbook to the response
        workbook.save(response)

        return response


    export_student_grades.short_description = 'Export Student Grades'

    actions = ['export_student_grades']

















class StudentAdmin(admin.ModelAdmin):
    list_display = ('name', 'school_class')
    list_filter = ('school_class',)


class TrimestreAdmin(admin.ModelAdmin):
    list_display = ('name',)

class GradeAdmin(admin.ModelAdmin):
    list_display = ('student', 'subject', 'trimestre', 'test')
    list_filter = ('trimestre', 'student__school_class__name')
    actions = ['export_grades']



class ArbiAdmin(admin.ModelAdmin):
    list_display = ('student', 'subject', 'trimestre', 'test', 'chifehi', 'kira2a', 'intej')
    list_filter = ('trimestre', 'student__school_class__name')


class FrancaiAdmin(admin.ModelAdmin):
    list_display = ('student', 'subject', 'trimestre', 'test', 'exp', 'lecture', 'production')
    list_filter = ('trimestre', 'student__school_class__name')
    

class AnglaiAdmin(admin.ModelAdmin):
    list_display = ('student', 'subject', 'trimestre', 'test', 'oral')
    list_filter = ('trimestre', 'student__school_class__name')
    

admin.site.register(Teacher)
  # Unregister the Grade model first
admin.site.register(Subject, SubjectAdmin)
admin.site.register(Student, StudentAdmin)
admin.site.register(SchoolClass, SchoolClassAdmin)
admin.site.register(Trimestre, TrimestreAdmin)
admin.site.register(Grade, GradeAdmin)  # Register the Grade model again with GradeAdmin
admin.site.register(Arbi, ArbiAdmin)
admin.site.register(Francai, FrancaiAdmin)
admin.site.register(Anglai, AnglaiAdmin)
